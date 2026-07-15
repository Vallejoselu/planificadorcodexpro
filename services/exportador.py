import csv
from datetime import UTC, datetime, time, timedelta
from html import escape
from pathlib import Path

from PySide6.QtGui import QPageSize, QPdfWriter, QTextDocument

from database.schema import DIAS_SEMANA
from repositories.calendario_repository import CalendarioRepository
from repositories.historial_repository import HistorialRepository
from repositories.repartidores_repository import RepartidoresRepository
from repositories.restaurantes_repository import RestaurantesRepository
from repositories.turnos_repository import TurnosRepository
from services.descansos import descanso_es_valido
from services.fechas import normalizar_fecha_inicio_semana


calendario_repository = CalendarioRepository()
repartidores_repository = RepartidoresRepository()
restaurantes_repository = RestaurantesRepository()
turnos_repository = TurnosRepository()
historial_repository = HistorialRepository()


def exportar_excel(ruta, fecha_inicio_semana=None):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    fecha_inicio_semana = normalizar_fecha_inicio_semana(fecha_inicio_semana)
    datos = preparar_datos_exportacion(fecha_inicio_semana)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Horarios"

    _escribir_hoja(
        hoja,
        datos["horarios"],
        ["Dia", "Turno", "Tipo", "Restaurante", "Zona", "Repartidor", "Inicio", "Fin", "Horas"]
    )
    _crear_hoja(
        libro,
        "Horas",
        datos["horas"],
        ["Concepto", "Nombre", "Horas"]
    )
    _crear_hoja(
        libro,
        "Descansos",
        datos["descansos"],
        ["Repartidor", "Horas contrato", "Descanso 1", "Descanso 2"]
    )
    _crear_hoja(
        libro,
        "Totales",
        datos["totales"],
        ["Concepto", "Valor"]
    )

    for hoja in libro.worksheets:

        for celda in hoja[1]:

            celda.font = Font(bold=True)
            celda.fill = PatternFill("solid", fgColor="D9EAF7")

        for columna in hoja.columns:

            ancho = max(len(str(celda.value or "")) for celda in columna)
            hoja.column_dimensions[columna[0].column_letter].width = ancho + 2

    libro.save(ruta)
    registrar_exportacion("Excel", ruta, fecha_inicio_semana)


def exportar_csv(ruta, fecha_inicio_semana=None):

    fecha_inicio_semana = normalizar_fecha_inicio_semana(fecha_inicio_semana)
    datos = preparar_datos_exportacion(fecha_inicio_semana)

    with open(ruta, "w", newline="", encoding="utf-8-sig") as archivo:

        escritor = csv.writer(archivo, delimiter=";")

        _escribir_bloque_csv(
            escritor,
            "Horarios",
            ["Dia", "Turno", "Tipo", "Restaurante", "Zona", "Repartidor", "Inicio", "Fin", "Horas"],
            datos["horarios"]
        )
        _escribir_bloque_csv(
            escritor,
            "Horas",
            ["Concepto", "Nombre", "Horas"],
            datos["horas"]
        )
        _escribir_bloque_csv(
            escritor,
            "Descansos",
            ["Repartidor", "Horas contrato", "Descanso 1", "Descanso 2"],
            datos["descansos"]
        )
        _escribir_bloque_csv(
            escritor,
            "Totales",
            ["Concepto", "Valor"],
            datos["totales"]
        )

    registrar_exportacion("CSV", ruta, fecha_inicio_semana)


def exportar_ics(ruta, fecha_inicio_semana=None):

    fecha_inicio_semana = normalizar_fecha_inicio_semana(fecha_inicio_semana)
    datos = preparar_datos_exportacion(fecha_inicio_semana)
    contenido = crear_calendario_ics(datos)

    with open(ruta, "w", newline="", encoding="utf-8") as archivo:

        archivo.write(contenido)

    registrar_exportacion("ICS", ruta, fecha_inicio_semana)


def exportar_pdf(ruta, fecha_inicio_semana=None):

    fecha_inicio_semana = normalizar_fecha_inicio_semana(fecha_inicio_semana)
    documento = QTextDocument()
    documento.setHtml(_crear_html(preparar_datos_exportacion(fecha_inicio_semana)))

    escritor = QPdfWriter(ruta)
    escritor.setPageSize(QPageSize(QPageSize.A4))
    escritor.setResolution(96)

    documento.print_(escritor)
    registrar_exportacion("PDF", ruta, fecha_inicio_semana)


def registrar_exportacion(formato, ruta, fecha_inicio_semana):

    historial_repository.registrar(
        "Exportar",
        "exportacion",
        f"{formato}: {ruta}",
        fecha_inicio_semana
    )


def preparar_datos_exportacion(fecha_inicio_semana=None):

    fecha_inicio_semana = normalizar_fecha_inicio_semana(
        fecha_inicio_semana
    )

    turnos = {
        turno[0]: turno
        for turno in turnos_repository.listar_todos()
    }
    restaurantes = {
        restaurante[0]: restaurante
        for restaurante in restaurantes_repository.listar_todos()
    }

    horarios = _preparar_horarios(
        turnos,
        restaurantes,
        fecha_inicio_semana
    )
    horas = _preparar_horas(horarios)
    descansos = _preparar_descansos()
    totales = _preparar_totales(horarios, descansos)

    return {
        "horarios": horarios,
        "horas": horas,
        "descansos": descansos,
        "totales": totales,
        "fecha_inicio_semana": fecha_inicio_semana
    }


def _preparar_horarios(turnos, restaurantes, fecha_inicio_semana):

    filas = []

    for asignacion in calendario_repository.listar_semana(fecha_inicio_semana):

        dia = asignacion[1]
        turno_id = asignacion[2]
        restaurante_id = asignacion[6]
        repartidor = asignacion[10] if len(asignacion) > 10 else ""
        turno = turnos.get(turno_id)
        restaurante = restaurantes.get(restaurante_id)

        if not turno or not restaurante:

            continue

        filas.append([
            dia,
            turno[2],
            turno[1],
            restaurante[1],
            restaurante[3] or "",
            repartidor or "",
            turno[3],
            turno[4],
            float(turno[6] or 0)
        ])

    filas.sort(key=lambda fila: (
        DIAS_SEMANA.index(fila[0]),
        fila[1],
        fila[3]
    ))

    return filas


def _preparar_horas(horarios):

    por_restaurante = {}
    por_turno = {}
    por_repartidor = {}

    for fila in horarios:

        restaurante = fila[3]
        turno = fila[1]
        repartidor = fila[5]
        horas = float(fila[8] or 0)

        por_restaurante[restaurante] = por_restaurante.get(restaurante, 0) + horas
        por_turno[turno] = por_turno.get(turno, 0) + horas

        if repartidor:

            por_repartidor[repartidor] = por_repartidor.get(repartidor, 0) + horas

    filas = []

    for restaurante, horas in sorted(por_restaurante.items()):

        filas.append(["Restaurante", restaurante, _formatear_numero(horas)])

    for turno, horas in sorted(por_turno.items()):

        filas.append(["Turno", turno, _formatear_numero(horas)])

    for repartidor, horas in sorted(por_repartidor.items()):

        filas.append(["Repartidor asignado", repartidor, _formatear_numero(horas)])

    for repartidor in repartidores_repository.listar_activos():

        filas.append([
            "Contrato repartidor",
            repartidor[1],
            repartidor[2]
        ])

    return filas


def _preparar_descansos():

    filas = []

    for repartidor in repartidores_repository.listar_activos():

        filas.append([
            repartidor[1],
            repartidor[2],
            _formatear_descanso_exportacion(repartidor, 9),
            _formatear_descanso_exportacion(repartidor, 10)
        ])

    return filas


def _formatear_descanso_exportacion(repartidor, posicion):

    if not repartidor[9] or not repartidor[10]:

        return repartidor[posicion] or ""

    if descanso_es_valido(repartidor[9], repartidor[10]):

        return repartidor[posicion]

    return f"{repartidor[posicion]} (no valido)"


def _preparar_totales(horarios, descansos):

    repartidores = repartidores_repository.listar_activos()
    total_horas = sum(float(fila[8] or 0) for fila in horarios)
    total_contratadas = sum(int(repartidor[2] or 0) for repartidor in repartidores)
    restaurantes = {
        fila[3]
        for fila in horarios
    }
    turnos = {
        fila[1]
        for fila in horarios
    }

    return [
        ["Total horarios", len(horarios)],
        ["Total horas planificadas", _formatear_numero(total_horas)],
        ["Total horas contratadas", total_contratadas],
        ["Total restaurantes", len(restaurantes)],
        ["Total tipos de turno", len(turnos)],
        ["Total repartidores", len(repartidores)],
        ["Total repartidores con descanso", len(descansos)]
    ]


def _crear_hoja(libro, nombre, filas, cabeceras):

    hoja = libro.create_sheet(nombre)
    _escribir_hoja(hoja, filas, cabeceras)


def _escribir_hoja(hoja, filas, cabeceras):

    hoja.append(cabeceras)

    for fila in filas:

        hoja.append(fila)


def _escribir_bloque_csv(escritor, titulo, cabeceras, filas):

    escritor.writerow([titulo])
    escritor.writerow(cabeceras)
    escritor.writerows(filas)
    escritor.writerow([])


def _crear_html(datos):

    partes = [
        "<html><body>",
        "<h1>Planificador Delivery Pro</h1>"
    ]

    bloques = [
        (
            "Horarios",
            ["Dia", "Turno", "Tipo", "Restaurante", "Zona", "Repartidor", "Inicio", "Fin", "Horas"],
            datos["horarios"]
        ),
        (
            "Horas",
            ["Concepto", "Nombre", "Horas"],
            datos["horas"]
        ),
        (
            "Descansos",
            ["Repartidor", "Horas contrato", "Descanso 1", "Descanso 2"],
            datos["descansos"]
        ),
        (
            "Totales",
            ["Concepto", "Valor"],
            datos["totales"]
        )
    ]

    for titulo, cabeceras, filas in bloques:

        partes.append(f"<h2>{titulo}</h2>")
        partes.append("<table border='1' cellspacing='0' cellpadding='4' width='100%'>")
        partes.append("<tr>")

        for cabecera in cabeceras:

            partes.append(f"<th>{escape(str(cabecera))}</th>")

        partes.append("</tr>")

        for fila in filas:

            partes.append("<tr>")

            for valor in fila:

                partes.append(f"<td>{escape(str(valor))}</td>")

            partes.append("</tr>")

        partes.append("</table>")

    partes.append("</body></html>")

    return "".join(partes)


def crear_calendario_ics(datos):

    fecha_inicio_semana = normalizar_fecha_inicio_semana(
        datos.get("fecha_inicio_semana")
    )
    eventos = []

    for indice, fila in enumerate(datos.get("horarios", []), start=1):

        evento = _crear_evento_ics(fila, fecha_inicio_semana, indice)

        if evento:

            eventos.extend(evento)

    lineas = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Planificador Delivery Pro//ES",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:Planificador Delivery Pro"
    ]
    lineas.extend(eventos)
    lineas.append("END:VCALENDAR")

    return "\r\n".join(_plegar_linea(linea) for linea in lineas) + "\r\n"


def _crear_evento_ics(fila, fecha_inicio_semana, indice):

    inicio = _datetime_evento(fecha_inicio_semana, fila[0], fila[6])
    fin = _datetime_evento(fecha_inicio_semana, fila[0], fila[7])

    if not inicio or not fin:

        return None

    if fin <= inicio:

        fin += timedelta(days=1)

    repartidor = fila[5] or "Sin repartidor"
    resumen = f"{fila[1]} - {fila[3]}"
    descripcion = (
        f"Tipo: {fila[2]}\n"
        f"Repartidor: {repartidor}\n"
        f"Zona: {fila[4] or ''}\n"
        f"Horas: {fila[8]}"
    )

    return [
        "BEGIN:VEVENT",
        f"UID:{_valor_ics(f'planificador-{fecha_inicio_semana}-{indice}@planificador-delivery-pro')}",
        f"DTSTAMP:{datetime.now(UTC).strftime('%Y%m%dT%H%M%SZ')}",
        f"DTSTART:{_formatear_datetime_ics(inicio)}",
        f"DTEND:{_formatear_datetime_ics(fin)}",
        f"SUMMARY:{_valor_ics(resumen)}",
        f"LOCATION:{_valor_ics(fila[3])}",
        f"DESCRIPTION:{_valor_ics(descripcion)}",
        "END:VEVENT"
    ]


def _datetime_evento(fecha_inicio_semana, dia, hora):

    if dia not in DIAS_SEMANA:

        return None

    hora = _parsear_hora(hora)

    if hora is None:

        return None

    fecha = (
        datetime.strptime(fecha_inicio_semana, "%Y-%m-%d").date()
        + timedelta(days=DIAS_SEMANA.index(dia))
    )

    return datetime.combine(fecha, hora)


def _parsear_hora(valor):

    valor = str(valor or "").strip()

    if not valor:

        return None

    for formato in ("%H:%M", "%H:%M:%S"):

        try:

            return datetime.strptime(valor, formato).time()

        except ValueError:

            continue

    return None


def _formatear_datetime_ics(valor):

    if isinstance(valor, time):

        raise TypeError("Se esperaba fecha y hora completas.")

    return valor.strftime("%Y%m%dT%H%M%S")


def _valor_ics(valor):

    texto = str(valor or "")
    texto = texto.replace("\\", "\\\\")
    texto = texto.replace(";", "\\;")
    texto = texto.replace(",", "\\,")
    texto = texto.replace("\r\n", "\\n")
    texto = texto.replace("\n", "\\n")

    return texto


def _plegar_linea(linea):

    partes = []
    restante = str(linea)

    while len(restante) > 75:

        partes.append(restante[:75])
        restante = " " + restante[75:]

    partes.append(restante)

    return "\r\n".join(partes)


def _formatear_numero(valor):

    if float(valor).is_integer():

        return int(valor)

    return round(valor, 2)


def normalizar_ruta(ruta, extension):

    ruta = Path(ruta)

    if ruta.suffix.lower() != extension:

        ruta = ruta.with_suffix(extension)

    return str(ruta)
