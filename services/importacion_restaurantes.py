import csv
from pathlib import Path

from repositories.ciudades_repository import CiudadesRepository
from repositories.historial_repository import HistorialRepository
from repositories.restaurantes_repository import RestaurantesRepository
from services.importacion_repartidores import (
    booleano,
    entero,
    normalizar_texto,
    texto
)


class ImportadorRestaurantes:

    CAMPOS = {
        "nombre": ("nombre", "restaurante", "local"),
        "ciudad": ("ciudad", "city"),
        "zona": ("zona", "area"),
        "direccion": ("direccion", "direccion postal", "address"),
        "telefono": ("telefono", "phone"),
        "prioridad": ("prioridad", "peso"),
        "observaciones": ("observaciones", "notas"),
        "activo": ("activo", "active"),
        "horario_comida": (
            "horario comida",
            "horario_comida",
            "comida"
        ),
        "horario_cena": (
            "horario cena",
            "horario_cena",
            "cena"
        )
    }

    def __init__(
        self,
        restaurantes_repository=None,
        ciudades_repository=None,
        historial_repository=None
    ):

        self.restaurantes_repository = (
            restaurantes_repository or RestaurantesRepository()
        )
        self.ciudades_repository = (
            ciudades_repository or CiudadesRepository()
        )
        self.historial_repository = (
            historial_repository or HistorialRepository()
        )

    def importar(self, ruta):

        ruta = Path(ruta)
        filas = self.leer_archivo(ruta)
        existentes = self.restaurantes_por_nombre()
        ciudades = self.ciudades_por_nombre()
        resultado = {
            "archivo": str(ruta),
            "leidos": len(filas),
            "creados": 0,
            "actualizados": 0,
            "errores": []
        }

        for numero, fila in enumerate(filas, start=2):

            try:

                datos = self.normalizar_fila(fila, ciudades)
                clave = normalizar_texto(datos["nombre"])
                existente = existentes.get(clave)

                if existente:

                    self.actualizar_restaurante(existente, datos)
                    resultado["actualizados"] += 1

                else:

                    restaurante_id = self.crear_restaurante(datos)
                    existentes[clave] = {
                        **datos,
                        "id": restaurante_id
                    }
                    resultado["creados"] += 1

            except ValueError as error:

                resultado["errores"].append({
                    "fila": numero,
                    "error": str(error)
                })

        self.registrar_historial(ruta, resultado)
        return resultado

    def leer_archivo(self, ruta):

        if not ruta.exists():

            raise ValueError("El archivo de importacion no existe.")

        extension = ruta.suffix.lower()

        if extension == ".csv":

            return self.leer_csv(ruta)

        if extension in (".xlsx", ".xlsm"):

            return self.leer_excel(ruta)

        raise ValueError("Formato no soportado. Usa CSV o Excel XLSX.")

    def leer_csv(self, ruta):

        contenido = ruta.read_text(encoding="utf-8-sig")
        muestra = contenido[:2048]

        try:

            dialecto = csv.Sniffer().sniff(muestra, delimiters=";,")

        except csv.Error:

            dialecto = csv.excel

        lector = csv.DictReader(contenido.splitlines(), dialect=dialecto)
        return [
            self.normalizar_claves(fila)
            for fila in lector
        ]

    def leer_excel(self, ruta):

        from openpyxl import load_workbook

        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        libro.close()

        if not filas:

            return []

        cabeceras = [
            normalizar_texto(celda)
            for celda in filas[0]
        ]
        datos = []

        for fila in filas[1:]:

            if not any(celda is not None and str(celda).strip() for celda in fila):

                continue

            datos.append({
                cabecera: valor
                for cabecera, valor in zip(cabeceras, fila)
            })

        return datos

    def normalizar_claves(self, fila):

        return {
            normalizar_texto(clave): valor
            for clave, valor in (fila or {}).items()
        }

    def normalizar_fila(self, fila, ciudades):

        campos_presentes = {
            campo
            for campo, alias in self.CAMPOS.items()
            if self.tiene_campo(fila, alias)
        }
        datos = {
            campo: self.valor(fila, alias)
            for campo, alias in self.CAMPOS.items()
        }
        nombre = texto(datos.get("nombre"))

        if not nombre:

            raise ValueError("El nombre es obligatorio.")

        ciudad_id = self.resolver_ciudad_id(datos.get("ciudad"), ciudades)

        return {
            "nombre": nombre,
            "ciudad_id": ciudad_id,
            "zona": texto(datos.get("zona")),
            "direccion": texto(datos.get("direccion")),
            "telefono": texto(datos.get("telefono")),
            "prioridad": entero(datos.get("prioridad"), "prioridad", 50),
            "observaciones": texto(datos.get("observaciones")),
            "activo": booleano(datos.get("activo"), True),
            "horario_comida": texto(datos.get("horario_comida")),
            "horario_cena": texto(datos.get("horario_cena")),
            "campos_presentes": campos_presentes
        }

    def valor(self, fila, alias):

        for nombre in alias:

            clave = normalizar_texto(nombre)

            if clave in fila:

                return fila[clave]

        return None

    def tiene_campo(self, fila, alias):

        for nombre in alias:

            if normalizar_texto(nombre) in fila:

                return True

        return False

    def restaurantes_por_nombre(self):

        resultado = {}

        for fila in self.restaurantes_repository.listar_todos():

            resultado[normalizar_texto(fila[1])] = {
                "id": fila[0],
                "nombre": fila[1]
            }

        return resultado

    def ciudades_por_nombre(self):

        resultado = {}

        for ciudad in self.ciudades_repository.listar_activas():

            resultado[normalizar_texto(ciudad[1])] = ciudad[0]

        return resultado

    def resolver_ciudad_id(self, ciudad, ciudades):

        nombre = texto(ciudad)

        if not nombre:

            return None

        clave = normalizar_texto(nombre)

        if clave not in ciudades:

            ciudades[clave] = self.ciudades_repository.crear(nombre)

        return ciudades[clave]

    def crear_restaurante(self, datos):

        return self.restaurantes_repository.crear(
            datos["nombre"],
            datos["direccion"],
            datos["zona"],
            datos["telefono"],
            datos["prioridad"],
            datos["observaciones"],
            int(datos["activo"]),
            datos["horario_comida"],
            datos["horario_cena"],
            ciudad_id=datos["ciudad_id"]
        )

    def actualizar_restaurante(self, existente, datos):

        restaurante_id = existente["id"]
        actual = self.restaurantes_repository.obtener_por_id(restaurante_id)
        campos_presentes = datos.get("campos_presentes", set())
        repartidores_fijos = (
            self.restaurantes_repository.obtener_repartidores_fijos(
                restaurante_id
            )
        )

        self.restaurantes_repository.actualizar(
            restaurante_id,
            datos["nombre"],
            self.valor_actualizado(datos, actual, "direccion", 2),
            self.valor_actualizado(datos, actual, "zona", 3),
            self.valor_actualizado(datos, actual, "telefono", 4),
            int(
                datos["activo"]
                if "activo" in campos_presentes
                else actual[6]
            ),
            self.valor_actualizado(datos, actual, "horario_comida", 7),
            self.valor_actualizado(datos, actual, "horario_cena", 8),
            repartidores_fijos=repartidores_fijos,
            ciudad_id=(
                datos["ciudad_id"]
                if "ciudad" in campos_presentes
                else actual[9]
            )
        )

    def valor_actualizado(self, datos, actual, campo, indice):

        if campo in datos.get("campos_presentes", set()):

            return datos[campo]

        return actual[indice]

    def registrar_historial(self, ruta, resultado):

        self.historial_repository.registrar(
            "Importar restaurantes",
            "restaurantes",
            (
                f"{ruta.name}: {resultado['creados']} creados, "
                f"{resultado['actualizados']} actualizados, "
                f"{len(resultado['errores'])} errores"
            )
        )
