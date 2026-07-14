import csv
from datetime import datetime
from pathlib import Path

from repositories.ausencias_repository import AusenciasRepository
from repositories.historial_repository import HistorialRepository
from repositories.repartidores_repository import RepartidoresRepository
from services.importacion_repartidores import (
    booleano,
    normalizar_texto,
    texto
)


class ImportadorAusencias:

    CAMPOS = {
        "nombre": ("nombre", "repartidor", "empleado"),
        "tipo": ("tipo", "ausencia", "clase"),
        "fecha_inicio": (
            "fecha inicio",
            "fecha_inicio",
            "inicio",
            "desde",
            "fecha"
        ),
        "fecha_fin": (
            "fecha fin",
            "fecha_fin",
            "fin",
            "hasta"
        ),
        "observaciones": ("observaciones", "notas"),
        "activo": ("activo", "activa", "vigente")
    }

    FORMATOS_FECHA = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d"
    )

    def __init__(
        self,
        repartidores_repository=None,
        ausencias_repository=None,
        historial_repository=None
    ):

        self.repartidores_repository = (
            repartidores_repository or RepartidoresRepository()
        )
        self.ausencias_repository = (
            ausencias_repository or AusenciasRepository()
        )
        self.historial_repository = (
            historial_repository or HistorialRepository()
        )

    def importar(self, ruta):

        ruta = Path(ruta)
        filas = self.leer_archivo(ruta)
        repartidores = self.repartidores_por_nombre()
        resultado = {
            "archivo": str(ruta),
            "leidos": len(filas),
            "vacaciones": 0,
            "bajas": 0,
            "duplicados": 0,
            "errores": []
        }

        for numero, fila in enumerate(filas, start=2):

            try:

                datos = self.normalizar_fila(fila, repartidores)
                creado = self.guardar_ausencia(datos)

                if creado:

                    resultado[datos["tipo"]] += 1

                else:

                    resultado["duplicados"] += 1

            except ValueError as error:

                resultado["errores"].append({
                    "fila": numero,
                    "error": str(error)
                })

        self.registrar_historial(ruta, resultado)
        return resultado

    def leer_archivo(self, ruta):

        if not ruta.exists():

            raise ValueError("El archivo de importacion no existe.")

        extension = ruta.suffix.lower()

        if extension == ".csv":

            return self.leer_csv(ruta)

        if extension in (".xlsx", ".xlsm"):

            return self.leer_excel(ruta)

        raise ValueError("Formato no soportado. Usa CSV o Excel XLSX.")

    def leer_csv(self, ruta):

        contenido = ruta.read_text(encoding="utf-8-sig")
        muestra = contenido[:2048]

        try:

            dialecto = csv.Sniffer().sniff(muestra, delimiters=";,")

        except csv.Error:

            dialecto = csv.excel

        lector = csv.DictReader(contenido.splitlines(), dialect=dialecto)
        return [
            self.normalizar_claves(fila)
            for fila in lector
        ]

    def leer_excel(self, ruta):

        from openpyxl import load_workbook

        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        libro.close()

        if not filas:

            return []

        cabeceras = [
            normalizar_texto(celda)
            for celda in filas[0]
        ]
        datos = []

        for fila in filas[1:]:

            if not any(celda is not None and str(celda).strip() for celda in fila):

                continue

            datos.append({
                cabecera: valor
                for cabecera, valor in zip(cabeceras, fila)
            })

        return datos

    def normalizar_claves(self, fila):

        return {
            normalizar_texto(clave): valor
            for clave, valor in (fila or {}).items()
        }

    def normalizar_fila(self, fila, repartidores):

        nombre = self.valor(fila, self.CAMPOS["nombre"])
        clave = normalizar_texto(nombre)

        if not clave:

            raise ValueError("El nombre es obligatorio.")

        repartidor = repartidores.get(clave)

        if not repartidor:

            raise ValueError(f"No existe el repartidor {texto(nombre)}.")

        tipo = self.normalizar_tipo(self.valor(fila, self.CAMPOS["tipo"]))
        fecha_inicio = self.normalizar_fecha(
            self.valor(fila, self.CAMPOS["fecha_inicio"]),
            "fecha_inicio"
        )
        fecha_fin = self.normalizar_fecha(
            self.valor(fila, self.CAMPOS["fecha_fin"]),
            "fecha_fin",
            requerida=False
        )

        if tipo == "vacaciones" and not fecha_fin:

            fecha_fin = fecha_inicio

        if fecha_fin and fecha_fin < fecha_inicio:

            raise ValueError("fecha_fin no puede ser anterior a fecha_inicio.")

        return {
            "repartidor_id": repartidor["id"],
            "tipo": tipo,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "observaciones": texto(self.valor(fila, self.CAMPOS["observaciones"])),
            "activo": booleano(self.valor(fila, self.CAMPOS["activo"]), True)
        }

    def normalizar_tipo(self, valor):

        tipo = normalizar_texto(valor)
        equivalencias = {
            "vacacion": "vacaciones",
            "vacaciones": "vacaciones",
            "holiday": "vacaciones",
            "baja": "bajas",
            "bajas": "bajas",
            "enfermedad": "bajas"
        }

        if tipo not in equivalencias:

            raise ValueError("Tipo no valido. Usa vacaciones o baja.")

        return equivalencias[tipo]

    def normalizar_fecha(self, valor, campo, requerida=True):

        if not texto(valor):

            if requerida:

                raise ValueError(f"{campo} es obligatoria.")

            return None

        if hasattr(valor, "date"):

            return valor.date().isoformat()

        if hasattr(valor, "isoformat") and not isinstance(valor, str):

            return valor.isoformat()

        valor_texto = texto(valor)

        for formato in self.FORMATOS_FECHA:

            try:

                return datetime.strptime(valor_texto, formato).date().isoformat()

            except ValueError:

                pass

        raise ValueError(f"{campo} debe tener una fecha valida.")

    def guardar_ausencia(self, datos):

        if datos["tipo"] == "vacaciones":

            _, creado = self.ausencias_repository.insertar_vacacion(
                datos["repartidor_id"],
                datos["fecha_inicio"],
                datos["fecha_fin"],
                datos["observaciones"],
                int(datos["activo"])
            )
            return creado

        _, creado = self.ausencias_repository.insertar_baja(
            datos["repartidor_id"],
            datos["fecha_inicio"],
            datos["fecha_fin"],
            datos["observaciones"],
            int(datos["activo"])
        )
        return creado

    def valor(self, fila, alias):

        for nombre in alias:

            clave = normalizar_texto(nombre)

            if clave in fila:

                return fila[clave]

        return None

    def repartidores_por_nombre(self):

        resultado = {}

        for fila in self.repartidores_repository.listar_activos():

            resultado[normalizar_texto(fila[1])] = {
                "id": fila[0],
                "nombre": fila[1]
            }

        return resultado

    def registrar_historial(self, ruta, resultado):

        self.historial_repository.registrar(
            "Importar vacaciones y bajas",
            "ausencias",
            (
                f"{ruta.name}: {resultado['vacaciones']} vacaciones, "
                f"{resultado['bajas']} bajas, "
                f"{resultado['duplicados']} duplicados, "
                f"{len(resultado['errores'])} errores"
            )
        )
