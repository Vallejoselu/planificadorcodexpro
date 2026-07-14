import csv
import unicodedata
from pathlib import Path

from database.schema import HORAS_CONTRATO
from repositories.historial_repository import HistorialRepository
from repositories.repartidores_repository import RepartidoresRepository


class ImportadorRepartidores:

    CAMPOS = {
        "nombre": ("nombre", "name"),
        "horas": ("horas", "horas contratadas", "contrato"),
        "zona": ("zona", "localidad"),
        "doble_turno": ("doble turno", "doble_turno"),
        "puede_hasta_la_una": (
            "puede hasta la una",
            "hasta la una",
            "puede_hasta_la_una"
        ),
        "prioridad_comida": ("prioridad comida", "prioridad_comida"),
        "prioridad_noche": ("prioridad noche", "prioridad_noche"),
        "prioridad_grela": ("prioridad grela", "prioridad_grela"),
        "observaciones": ("observaciones", "notas"),
        "descanso_inicio": ("descanso inicio", "descanso_inicio"),
        "descanso_fin": ("descanso fin", "descanso_fin"),
        "apoyo_flexible": ("apoyo flexible", "apoyo_flexible"),
        "horas_complementarias": (
            "horas complementarias",
            "horas_complementarias"
        ),
        "max_horas_diarias": ("max horas diarias", "max_horas_diarias"),
        "max_dias_consecutivos": (
            "max dias consecutivos",
            "max_dias_consecutivos"
        )
    }

    def __init__(
        self,
        repartidores_repository=None,
        historial_repository=None
    ):

        self.repartidores_repository = (
            repartidores_repository or RepartidoresRepository()
        )
        self.historial_repository = (
            historial_repository or HistorialRepository()
        )

    def importar(self, ruta):

        ruta = Path(ruta)
        filas = self.leer_archivo(ruta)
        existentes = self.repartidores_por_nombre()
        resultado = {
            "archivo": str(ruta),
            "leidos": len(filas),
            "creados": 0,
            "actualizados": 0,
            "errores": []
        }

        for numero, fila in enumerate(filas, start=2):

            try:

                datos = self.normalizar_fila(fila)
                clave = normalizar_texto(datos["nombre"])
                existente = existentes.get(clave)

                if existente:

                    self.actualizar_repartidor(existente, datos)
                    resultado["actualizados"] += 1

                else:

                    repartidor_id = self.crear_repartidor(datos)
                    existentes[clave] = {
                        **datos,
                        "id": repartidor_id
                    }
                    resultado["creados"] += 1

            except ValueError as error:

                resultado["errores"].append({
                    "fila": numero,
                    "error": str(error)
                })

        self.registrar_historial(ruta, resultado)
        return resultado

    def leer_archivo(self, ruta):

        if not ruta.exists():

            raise ValueError("El archivo de importacion no existe.")

        extension = ruta.suffix.lower()

        if extension == ".csv":

            return self.leer_csv(ruta)

        if extension in (".xlsx", ".xlsm"):

            return self.leer_excel(ruta)

        raise ValueError("Formato no soportado. Usa CSV o Excel XLSX.")

    def leer_csv(self, ruta):

        texto = ruta.read_text(encoding="utf-8-sig")
        muestra = texto[:2048]

        try:

            dialecto = csv.Sniffer().sniff(muestra, delimiters=";,")

        except csv.Error:

            dialecto = csv.excel

        lector = csv.DictReader(texto.splitlines(), dialect=dialecto)
        return [
            self.normalizar_claves(fila)
            for fila in lector
        ]

    def leer_excel(self, ruta):

        from openpyxl import load_workbook

        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        libro.close()

        if not filas:

            return []

        cabeceras = [
            normalizar_texto(celda)
            for celda in filas[0]
        ]
        datos = []

        for fila in filas[1:]:

            if not any(celda is not None and str(celda).strip() for celda in fila):

                continue

            datos.append({
                cabecera: valor
                for cabecera, valor in zip(cabeceras, fila)
            })

        return datos

    def normalizar_claves(self, fila):

        return {
            normalizar_texto(clave): valor
            for clave, valor in (fila or {}).items()
        }

    def normalizar_fila(self, fila):

        datos = {
            campo: self.valor(fila, alias)
            for campo, alias in self.CAMPOS.items()
        }
        nombre = texto(datos.get("nombre"))

        if not nombre:

            raise ValueError("El nombre es obligatorio.")

        horas = entero(datos.get("horas"), "horas")

        if horas not in HORAS_CONTRATO:

            raise ValueError(
                "Horas contratadas no validas. Usa 10, 20, 25, 30, 35 o 40."
            )

        return {
            "nombre": nombre,
            "horas": horas,
            "zona": texto(datos.get("zona")),
            "doble_turno": booleano(datos.get("doble_turno"), True),
            "puede_hasta_la_una": booleano(
                datos.get("puede_hasta_la_una"),
                True
            ),
            "prioridad_comida": entero(
                datos.get("prioridad_comida"),
                "prioridad_comida",
                50
            ),
            "prioridad_noche": entero(
                datos.get("prioridad_noche"),
                "prioridad_noche",
                50
            ),
            "prioridad_grela": entero(
                datos.get("prioridad_grela"),
                "prioridad_grela",
                50
            ),
            "observaciones": texto(datos.get("observaciones")),
            "descanso_inicio": texto(datos.get("descanso_inicio")) or None,
            "descanso_fin": texto(datos.get("descanso_fin")) or None,
            "apoyo_flexible": booleano(datos.get("apoyo_flexible"), False),
            "horas_complementarias": entero(
                datos.get("horas_complementarias"),
                "horas_complementarias",
                0
            ),
            "max_horas_diarias": entero(
                datos.get("max_horas_diarias"),
                "max_horas_diarias",
                10
            ),
            "max_dias_consecutivos": entero(
                datos.get("max_dias_consecutivos"),
                "max_dias_consecutivos",
                5
            )
        }

    def valor(self, fila, alias):

        for nombre in alias:

            clave = normalizar_texto(nombre)

            if clave in fila:

                return fila[clave]

        return None

    def repartidores_por_nombre(self):

        resultado = {}

        for fila in self.repartidores_repository.listar_activos():

            if isinstance(fila, dict):

                resultado[normalizar_texto(fila.get("nombre"))] = fila

            else:

                resultado[normalizar_texto(fila[1])] = {
                    "id": fila[0],
                    "nombre": fila[1]
                }

        return resultado

    def crear_repartidor(self, datos):

        return self.repartidores_repository.crear(
            datos["nombre"],
            datos["horas"],
            datos["zona"],
            int(datos["doble_turno"]),
            int(datos["puede_hasta_la_una"]),
            datos["prioridad_comida"],
            datos["prioridad_noche"],
            datos["prioridad_grela"],
            datos["observaciones"],
            descanso_inicio=datos["descanso_inicio"],
            descanso_fin=datos["descanso_fin"],
            apoyo_flexible=int(datos["apoyo_flexible"]),
            horas_complementarias=datos["horas_complementarias"],
            max_horas_diarias=datos["max_horas_diarias"],
            max_dias_consecutivos=datos["max_dias_consecutivos"]
        )

    def actualizar_repartidor(self, existente, datos):

        repartidor_id = existente["id"]
        actual = self.repartidores_repository.obtener_por_id(repartidor_id)

        if not actual:

            raise ValueError(f"No se encontro {datos['nombre']} para actualizar.")

        self.repartidores_repository.actualizar(
            repartidor_id,
            datos["nombre"],
            datos["horas"],
            datos["zona"],
            int(datos["doble_turno"]),
            int(datos["puede_hasta_la_una"]),
            datos["prioridad_comida"],
            datos["prioridad_noche"],
            datos["prioridad_grela"],
            datos["observaciones"],
            descanso_inicio=datos["descanso_inicio"],
            descanso_fin=datos["descanso_fin"],
            disponibilidad=actual.get("disponibilidad"),
            ciudad_principal_id=actual.get("ciudad_principal_id"),
            restaurante_principal_id=actual.get("restaurante_principal_id"),
            apoyo_flexible=int(datos["apoyo_flexible"]),
            horas_complementarias=datos["horas_complementarias"],
            max_horas_diarias=datos["max_horas_diarias"],
            max_dias_consecutivos=datos["max_dias_consecutivos"],
            ciudades_autorizadas=actual.get("ciudades_autorizadas"),
            restaurantes_autorizados=actual.get("restaurantes_autorizados")
        )

    def registrar_historial(self, ruta, resultado):

        self.historial_repository.registrar(
            "Importar repartidores",
            "repartidores",
            (
                f"{ruta.name}: {resultado['creados']} creados, "
                f"{resultado['actualizados']} actualizados, "
                f"{len(resultado['errores'])} errores"
            )
        )


def normalizar_texto(valor):

    texto_valor = texto(valor).lower()
    texto_valor = unicodedata.normalize("NFKD", texto_valor)
    texto_valor = "".join(
        caracter
        for caracter in texto_valor
        if not unicodedata.combining(caracter)
    )
    return " ".join(texto_valor.replace("_", " ").split())


def texto(valor):

    if valor is None:

        return ""

    return str(valor).strip()


def entero(valor, campo, defecto=None):

    if valor is None or texto(valor) == "":

        if defecto is not None:

            return defecto

        raise ValueError(f"{campo} es obligatorio.")

    try:

        return int(float(str(valor).replace(",", ".")))

    except ValueError as error:

        raise ValueError(f"{campo} debe ser numerico.") from error


def booleano(valor, defecto=False):

    if valor is None or texto(valor) == "":

        return bool(defecto)

    valor_normalizado = normalizar_texto(valor)

    if valor_normalizado in ("1", "si", "s", "true", "verdadero", "yes"):

        return True

    if valor_normalizado in ("0", "no", "n", "false", "falso"):

        return False

    return bool(defecto)
