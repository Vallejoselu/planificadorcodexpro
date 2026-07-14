import csv
from pathlib import Path

from database.schema import DIAS_SEMANA, OPCIONES_DISPONIBILIDAD
from repositories.historial_repository import HistorialRepository
from repositories.repartidores_repository import RepartidoresRepository
from services.importacion_repartidores import (
    booleano,
    normalizar_texto,
    texto
)


class ImportadorDisponibilidad:

    CAMPOS_BASE = {
        "nombre": ("nombre", "repartidor", "empleado"),
        "dia": ("dia", "dia semana", "dia_semana"),
        "disponibilidad": ("disponibilidad", "opcion", "estado"),
        "turno": ("turno", "franja"),
        "disponible": ("disponible", "trabaja")
    }

    def __init__(
        self,
        repartidores_repository=None,
        historial_repository=None
    ):

        self.repartidores_repository = (
            repartidores_repository or RepartidoresRepository()
        )
        self.historial_repository = (
            historial_repository or HistorialRepository()
        )

    def importar(self, ruta):

        ruta = Path(ruta)
        filas = self.leer_archivo(ruta)
        repartidores = self.repartidores_por_nombre()
        cambios = {}
        resultado = {
            "archivo": str(ruta),
            "leidos": len(filas),
            "actualizados": 0,
            "errores": []
        }

        for numero, fila in enumerate(filas, start=2):

            try:

                nombre = self.valor(fila, self.CAMPOS_BASE["nombre"])
                clave = normalizar_texto(nombre)

                if not clave:

                    raise ValueError("El nombre es obligatorio.")

                repartidor = repartidores.get(clave)

                if not repartidor:

                    raise ValueError(f"No existe el repartidor {texto(nombre)}.")

                disponibilidad = self.disponibilidad_de_fila(fila)

                if not disponibilidad:

                    raise ValueError("No hay disponibilidad para importar.")

                destino = cambios.setdefault(
                    repartidor["id"],
                    self.disponibilidad_actual(repartidor["id"])
                )
                destino.update(disponibilidad)

            except ValueError as error:

                resultado["errores"].append({
                    "fila": numero,
                    "error": str(error)
                })

        for repartidor_id, disponibilidad in cambios.items():

            self.repartidores_repository.guardar_disponibilidad(
                repartidor_id,
                disponibilidad
            )
            resultado["actualizados"] += 1

        self.registrar_historial(ruta, resultado)
        return resultado

    def leer_archivo(self, ruta):

        if not ruta.exists():

            raise ValueError("El archivo de importacion no existe.")

        extension = ruta.suffix.lower()

        if extension == ".csv":

            return self.leer_csv(ruta)

        if extension in (".xlsx", ".xlsm"):

            return self.leer_excel(ruta)

        raise ValueError("Formato no soportado. Usa CSV o Excel XLSX.")

    def leer_csv(self, ruta):

        contenido = ruta.read_text(encoding="utf-8-sig")
        muestra = contenido[:2048]

        try:

            dialecto = csv.Sniffer().sniff(muestra, delimiters=";,")

        except csv.Error:

            dialecto = csv.excel

        lector = csv.DictReader(contenido.splitlines(), dialect=dialecto)
        return [
            self.normalizar_claves(fila)
            for fila in lector
        ]

    def leer_excel(self, ruta):

        from openpyxl import load_workbook

        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        libro.close()

        if not filas:

            return []

        cabeceras = [
            normalizar_texto(celda)
            for celda in filas[0]
        ]
        datos = []

        for fila in filas[1:]:

            if not any(celda is not None and str(celda).strip() for celda in fila):

                continue

            datos.append({
                cabecera: valor
                for cabecera, valor in zip(cabeceras, fila)
            })

        return datos

    def normalizar_claves(self, fila):

        return {
            normalizar_texto(clave): valor
            for clave, valor in (fila or {}).items()
        }

    def disponibilidad_de_fila(self, fila):

        dia = self.valor(fila, self.CAMPOS_BASE["dia"])

        if texto(dia):

            return self.disponibilidad_larga(fila)

        return self.disponibilidad_semanal(fila)

    def disponibilidad_larga(self, fila):

        dia = self.normalizar_dia(self.valor(fila, self.CAMPOS_BASE["dia"]))
        opcion = self.valor(fila, self.CAMPOS_BASE["disponibilidad"])

        if texto(opcion):

            return {
                dia: self.normalizar_opcion(opcion)
            }

        turno = self.valor(fila, self.CAMPOS_BASE["turno"])
        disponible = self.valor(fila, self.CAMPOS_BASE["disponible"])

        if not texto(turno):

            raise ValueError("Indica disponibilidad o turno.")

        if not booleano(disponible, True):

            return {
                dia: "No disponible"
            }

        return {
            dia: self.opcion_desde_turno(turno)
        }

    def disponibilidad_semanal(self, fila):

        disponibilidad = {}

        for dia in DIAS_SEMANA:

            if dia not in fila:

                continue

            if not texto(fila[dia]):

                continue

            disponibilidad[dia] = self.normalizar_opcion(fila[dia])

        return disponibilidad

    def normalizar_dia(self, valor):

        dia = normalizar_texto(valor)

        if dia not in DIAS_SEMANA:

            raise ValueError("Dia de disponibilidad no valido.")

        return dia

    def normalizar_opcion(self, valor):

        opcion = normalizar_texto(valor)

        equivalencias = {
            "comida": "Comidas",
            "comidas": "Comidas",
            "cena": "Cenas",
            "cenas": "Cenas",
            "noche": "Cenas",
            "noches": "Cenas",
            "ambos": "Ambos",
            "ambas": "Ambos",
            "si": "Ambos",
            "disponible": "Ambos",
            "todo": "Ambos",
            "no": "No disponible",
            "ninguno": "No disponible",
            "libre": "No disponible",
            "no disponible": "No disponible"
        }

        if opcion not in equivalencias:

            raise ValueError(
                "Disponibilidad no valida. Usa Comidas, Cenas, Ambos o No disponible."
            )

        valor_normalizado = equivalencias[opcion]

        if valor_normalizado not in OPCIONES_DISPONIBILIDAD:

            raise ValueError("Disponibilidad no valida.")

        return valor_normalizado

    def opcion_desde_turno(self, turno):

        turno = normalizar_texto(turno)

        if turno in ("comida", "comidas"):

            return "Comidas"

        if turno in ("cena", "cenas", "noche", "noches"):

            return "Cenas"

        if turno in ("ambos", "ambas"):

            return "Ambos"

        raise ValueError("Turno de disponibilidad no valido.")

    def valor(self, fila, alias):

        for nombre in alias:

            clave = normalizar_texto(nombre)

            if clave in fila:

                return fila[clave]

        return None

    def repartidores_por_nombre(self):

        resultado = {}

        for fila in self.repartidores_repository.listar_activos():

            resultado[normalizar_texto(fila[1])] = {
                "id": fila[0],
                "nombre": fila[1]
            }

        return resultado

    def disponibilidad_actual(self, repartidor_id):

        repartidor = self.repartidores_repository.obtener_por_id(repartidor_id)
        disponibilidad = {
            dia: "Ambos"
            for dia in DIAS_SEMANA
        }

        for dia, turnos in (repartidor.get("disponibilidad") or {}).items():

            disponibilidad[dia] = self.opcion_desde_turnos(turnos)

        return disponibilidad

    def opcion_desde_turnos(self, turnos):

        turnos = turnos or []

        if "comida" in turnos and "noche" in turnos:

            return "Ambos"

        if "comida" in turnos:

            return "Comidas"

        if "noche" in turnos:

            return "Cenas"

        return "No disponible"

    def registrar_historial(self, ruta, resultado):

        self.historial_repository.registrar(
            "Importar disponibilidad",
            "disponibilidad",
            (
                f"{ruta.name}: {resultado['actualizados']} repartidores "
                f"actualizados, {len(resultado['errores'])} errores"
            )
        )
